'''
IMPOTACIONES
'''
# Archivos
import openpyxl

# Correos
import smtplib, ssl
import getpass
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.message import EmailMessage

# GUI
import tkinter as tk
from tkinter import ttk, CENTER
from tkinter import messagebox
from tkinter import filedialog

class Envio:

    def __init__(self, ventana):
        # Variables
        self.ruta_archivo = ''
        self.empleados_procesados = []

        '''
        INICIO GUI
        '''

        # Ventana principal 
        self.ventana_principal = ventana
        self.ventana_principal.title('Sistema de correos')

        # Ventanas inicien en el centro de la pantalla
        self.windowWidth = self.ventana_principal.winfo_reqwidth()
        self.windowHeight = self.ventana_principal.winfo_reqheight()
        self.positionRight = int(self.ventana_principal.winfo_screenwidth()/2 - self.windowWidth/2)
        self.positionDown = int(self.ventana_principal.winfo_screenheight()/2 - self.windowHeight/2)

        # Campos de la ventana principal

        self.mensaje_bienvenida =tk.Label( self.ventana_principal, text = '¡Bienvenido al Envia Correos Automatizados!', font=( "Helvetica", 14 ) )

        self.titular_estructura = tk.Label( self.ventana_principal, text = "El excel a seleccionar debe de tener la siguiente estructura:", font=( "Helvetica", 12 ) )

        self.titular_estructura_excel = tk.Label( self.ventana_principal, text = "|NOMBRE||APELLIDO||CORREO||SALARIO BASE||DESCUENTOS||BONOS|", font=( "Helvetica", 10 ) )

        self.titular_archivo = tk.Label( self.ventana_principal, text = "Seleccionar archivo" )

        self.boton_abrir_archivo = tk.Button( self.ventana_principal, text="Abrir archivo", command = self.seleccionar_archivo )

        self.archivo_seleccionado = tk.Label( self.ventana_principal, text = "", fg="green" )

        self.boton_limpiar_tabla = tk.Button( self.ventana_principal, text="Limpiar tabla", command = self.boton_limpiar_tabla )

        self.mensaje_contenido_excel = tk.Label( self.ventana_principal, text = 'Verifica los correos y los salarios totales antes de enviar los correos. El programa calculará automáticamente el Salario Total.', fg="gray" ) 

        self.tree = ttk.Treeview()
        
        self.boton_abrir_ventana_login = tk.Button( self.ventana_principal, text="Enviar correos", command = self.abrir_ventana_login, state=tk.DISABLED, font=( "Helvetica", 14 ) )

        # Alineación de campos ventana principal

        self.mensaje_bienvenida.grid( row = 0, columnspan = 2, pady = '5' )
        self.titular_estructura.grid( row = 1, columnspan = 2, pady = '5' )
        self.titular_estructura_excel.grid( row = 2, columnspan = 2, pady = '5' )
        self.titular_archivo.grid( row = 3, column = 0 )
        self.boton_abrir_archivo.grid( row = 3, column = 1 )
        self.archivo_seleccionado.grid( row = 4, columnspan = 2, pady = '5' )
        self.mensaje_contenido_excel.grid( row = 5, columnspan = 2, pady = '5' )
        self.boton_limpiar_tabla.grid( row = 6, columnspan = 2, pady = '5' )
        self.tree.grid( row = 7, columnspan = 2, pady = '10' )
        self.boton_abrir_ventana_login.grid( row = 8, columnspan = 2, sticky='nesw', ipady = '5' )

        # Configuración tabla

        self.tree['columns'] = ('nombre','apellido','correo','salario_base', 'descuentos', 'bonos', 'salario_total')
        self.tree.column('nombre', anchor=CENTER, width=100, minwidth=100, stretch=tk.YES)
        self.tree.column('apellido', anchor=CENTER, width=100, minwidth=100, stretch=tk.YES)
        self.tree.column('correo', anchor=CENTER, width=200, minwidth=200, stretch=tk.YES)
        self.tree.column('salario_base', anchor=CENTER, width=100, minwidth=100, stretch=tk.YES)
        self.tree.column('descuentos', anchor=CENTER, width=100, minwidth=100, stretch=tk.YES)
        self.tree.column('bonos', anchor=CENTER, width=100, minwidth=100, stretch=tk.YES)
        self.tree.column('salario_total', anchor=CENTER, width=100, minwidth=100, stretch=tk.NO)

        self.tree.heading('#0', text="ID", anchor=CENTER)
        self.tree.heading('#1', text="NOMBRE", anchor=CENTER)
        self.tree.heading('#2', text="APELLIDO", anchor=CENTER)
        self.tree.heading('#3', text="CORREO", anchor=CENTER)
        self.tree.heading('#4', text="SALARIO BASE", anchor=CENTER)
        self.tree.heading('#5', text="DESCUENTOS", anchor=CENTER)
        self.tree.heading('#6', text="BONOS", anchor=CENTER)
        self.tree.heading('#7', text="SALARIO TOTAL", anchor=CENTER)

        # Ventana de login
        self.ventana_login = tk.Toplevel( self.ventana_principal )
        self.ventana_login.title('Login')
        self.ventana_login.withdraw()
        self.ventana_login.geometry("+{}+{}".format(self.positionRight, self.positionDown + 100))
        # Prevenir la opción de cerrar la ventana
        self.ventana_login.protocol("WM_DELETE_WINDOW", self.disable_event)
            
        # Campos de la ventana de login

        self.titular_registro = tk.Label( self.ventana_login, text = "Inicia sesión para poder enviar los correos." )

        self.titular_correo = tk.Label( self.ventana_login, text = "Correo" )
        self.entrada_correo = tk.Entry( self.ventana_login )

        self.titular_password = tk.Label( self.ventana_login, text = "Contraseña" )
        self.entrada_password = tk.Entry( self.ventana_login, show="*" )

        self.boton_inicia_sesion_enviar_correos = tk.Button( self.ventana_login, text="Enviar correos", command = self.proceso_envio_resultados )

        self.boton_salir_ventana_login = tk.Button( self.ventana_login, text="Salir", command = self.salir_ventana_login ) 

        # Alineación campos ventana login

        self.titular_registro.grid( row = 0, columnspan = 2, pady = '5', padx = '5' )
        self.titular_correo.grid( row = 1, column = 0, pady = '5', padx = '5' )
        self.entrada_correo.grid( row = 1, column = 1, pady = '5', padx = '5' )
        self.titular_password.grid( row = 2, column = 0, pady = '5', padx = '5' )
        self.entrada_password.grid( row = 2, column = 1, pady = '5', padx = '5' )
        self.boton_salir_ventana_login.grid( row = 3, column = 0, sticky='nesw', pady = '5', padx = '5' )    
        self.boton_inicia_sesion_enviar_correos.grid( row = 3, column = 1, sticky='nesw', pady = '5', padx = '5' )  

        '''
        FIN GUI
        '''
    
    # Abrir ventana de login
    def abrir_ventana_login( self ):
        # Muestra la ventana login 
        self.ventana_login.deiconify()
        
    # Proceso envio de correos
    def proceso_envio_resultados( self ):
        
        try:
            # Enviar los correos
            self.envio_correos( self.empleados_procesados )
            # Genera un archivo de resultados xlsx
            self.escribir_archivo_resultados( self.empleados_procesados )
            # Salir ventana login
            self.salir_ventana_login()
            # Alertas completados exitosamente
            messagebox.showinfo('Envio exitoso', 'Los correos se han enviado exitosamente.')
            messagebox.showinfo('Archivo de resultados', 'Se ha creado un archivo llamado resultados.xlsx.')

        except EnvironmentError:
            # Alerta de inicio de sesión fallida
            messagebox.showerror( 'Error login', 'Por favor verifica tu usuario y contraseña.' )

    # Selecciona un archivo en el equipo
    def seleccionar_archivo( self ):
        # Obtener la ruta del archivo
        self.ruta_archivo = tk.filedialog.askopenfilename(title="Abrir archivo excel")

        # Verificar si hay algún archivo seleccionado
        if( self.ruta_archivo ):
            # Setea el valor de la ruta en una etiqueta de la ventana
            self.archivo_seleccionado['text'] = self.ruta_archivo
            # Setea el resultado que obtiene de la funcion
            self.empleados_procesados = self.obtener_empleados_procesados( self.ruta_archivo )
            # Muestra el resultado en la tabla
            self.mostrar_en_tabla( self.empleados_procesados )
            # Habilita el boton de envio
            self.habilitar_boton_principal()

            print( self.empleados_procesados )
        else:
            # Alerta de que no se ha seleccionado un archivo
            tk.messagebox.showwarning('Selecciona un archivo', 'Por favor selecciona un archivo de excel con la estructura antes mencionada.')
    
    # Funcion para limpiar y llenar la tabla
    def mostrar_en_tabla(self, lista):
        # Limpia el listado para ingresar los nuevos valores
        self.limpiar_listado_empleados()
        # Setea la lista recibida a la variable
        self.empleados_procesados = lista
        # Limpia la tabla
        self.limpiar_tabla()
        # Llena la tabla con los nuevos datos
        for empleado in lista:
            self.tree.insert('', 0, text= empleado['id'], values = (
                empleado['nombre'], 
                empleado['apellido'], 
                empleado['correo'], 
                empleado['salario_base'], 
                empleado['descuentos'], 
                empleado['bonos'],
                empleado['salario_total']
                )
            )

    # Limpia los datos de la tabla 
    def limpiar_tabla( self ):
        registros = self.tree.get_children()
        for element in registros:
            self.tree.delete(element)

    # Limpia el listado general
    def limpiar_listado_empleados( self ):
        self.empleados_procesados = []
    
    # Limpia todos los campos con datos
    def boton_limpiar_tabla( self ):
        self.limpiar_listado_empleados()
        self.limpiar_tabla()
        self.deshabilitar_boton_principal()
        self.archivo_seleccionado['text'] = ""
        messagebox.showinfo( 'Registros limpiados', 'Los registros se han limpiado correctamente.' )

    # Habilita el botón de envio
    def habilitar_boton_principal( self ):
        self.boton_abrir_ventana_login['state'] = tk.NORMAL

    # Deshabilita el botón de envio
    def deshabilitar_boton_principal( self ):
        self.boton_abrir_ventana_login['state'] = tk.DISABLED

    # Función que impide el cierre de la ventana login
    def disable_event( self ):
        pass

    # Salir de la ventana login
    def salir_ventana_login( self ):
        self.ventana_login.withdraw()

        
    '''
    INICIO PROCESO ENVIO DE CORREOS
    '''

    def envio_correos( self, empleados_procesados ):

        # Pedir datos para inicio de sesión
        username = self.entrada_correo.get() 
        password = self.entrada_password.get()

        # Crear la conexión
        context = ssl.create_default_context()

        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as server:
            # Inicio de sesion
            server.login( username, password )
            
            print( '-> Inicio de sesion exitoso.\n' )

            for empleado_procesado in empleados_procesados:

                # Destinatario
                destinatario = empleado_procesado['correo']

                # Crear el mensaje
                msg = EmailMessage()
                msg.set_content('Hola {} {}. Este mes recibiste un salario total de Q. {}.'.format( empleado_procesado['nombre'], empleado_procesado['apellido'], empleado_procesado['salario_total'] ) )
                msg['Subject'] = '¡Saludos {}, tu salario ha sido depositado!'.format( empleado_procesado['nombre'] )
                msg['From'] = username
                msg['To'] = destinatario

                # Enviar los mensajes
                server.send_message(msg)
                print( '-> Mensaje enviado al correo: {}'.format( empleado_procesado['correo'] ) )

        print("\n-> Proceso terminado, envio de correos exitoso. =D\n")
        

    '''
    FIN PROCESO ENVIO DE CORREOS
    '''


    '''
    INICIO PROCESO LEER ARCHIVO
    '''


    def obtener_empleados_procesados( self, ruta_archivo ):

        # Leer el archivo
        book = openpyxl.load_workbook(ruta_archivo, read_only=True)

        # Fijar la hoja
        hoja = book.active

        # Listado de empleados procesados
        empleados_procesados = []

        # Contador inicial de lectura
        contador_lectura = 0

        # Obtener salario total
        def obtener_salario_total(salario_base, descuentos, bonos):
            resultado = salario_base - descuentos + bonos
            return resultado

        # Obtiene los datos del excel y
        for fila in hoja.iter_rows(min_col=1, max_col=hoja.max_column):

            # Comprension de listas
            empleado = [celda.value for celda in fila]

            # Quita la primera fila
            if(contador_lectura > 0):
            # Valida si viene el correo el eletrónico
                if(empleado[2] is not None):
                    # Obtiene los datos para calcular el salario total
                    salario_base = int(empleado[3])
                    descuentos = int(empleado[4])
                    bonos = int(empleado[5])

                    # Obtiene el salario total
                    salario_total = obtener_salario_total(salario_base, descuentos, bonos)

                    # Se crea el diccionario empleado procesado
                    empleado_procesado = {
                        "id": contador_lectura,
                        "nombre": empleado[0],
                        "apellido": empleado[1],
                        "correo": empleado[2],
                        "salario_base": empleado[3],
                        "descuentos": empleado[4],
                        "bonos": empleado[5],
                        "salario_total": salario_total
                    }

                    # Se agrega el diccionario al listado de empleados procesados
                    empleados_procesados.append( empleado_procesado )

            # Incrementa el contador de lectura
            contador_lectura = contador_lectura + 1
        
        # Retornar el listado de listados procesados 
        return empleados_procesados
        print( '-> Datos procesados correctamente.' )



    '''
    FIN PROCESO LEER ARCHIVO
    '''

    '''
    INICIO PROCESO ESCRIBIR ARCHIVO
    '''


    def escribir_archivo_resultados( self, empleados_procesados ):
        # Abre un libro
        libro_resultados = openpyxl.Workbook()

        # Fijar la hoja
        hoja_activa = libro_resultados.active

        # Se agregan los encabezados
        hoja_activa.append([
            'Nombre', 
            'Apellido', 
            'Correo', 
            'Salario base', 
            'Descuentos', 
            'Bonos', 
            'Salario Total'
        ])

        # Se agregan los datos en el archivo de resultados
        for empleado_procesado in empleados_procesados:
            hoja_activa.append(
                [
                    empleado_procesado['nombre'],
                    empleado_procesado['apellido'],
                    empleado_procesado['correo'],
                    empleado_procesado['salario_base'],
                    empleado_procesado['descuentos'],
                    empleado_procesado['bonos'],
                    empleado_procesado['salario_total']
                ]
            )

        # Se crea el archivo
        libro_resultados.save('resultados.xlsx')
        print( '-> Se ha creado un registro con el nombre de resultados.xlsx.' )


    '''
    FIN PROCESO ESCRIBIR ARCHIVO
    '''

'''
    INICIO DE LA APLICACIÓN
'''

if __name__ == '__main__':
    ventana = tk.Tk()
    aplicacion = Envio( ventana )
    ventana.mainloop()